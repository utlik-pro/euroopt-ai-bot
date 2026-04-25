"""Парсер xlsx-справочника магазинов в data/stores_new/.

Источник: «Список ТО ЕТ Хит с форматами.xlsx» — внутренний реестр Евроторга
с правильной разметкой бренда (Хит / Евроопт), формата (Маркет / Супер /
Гипер / Хит Стандарт / Хит-экспресс / Автолавка / ...) и полным адресом.

Парсер:
1. Читает xlsx через openpyxl.
2. Извлекает из строки адреса город, улицу, дом (строки вида
   «33 Магазин г.Минск,ул.Рафиева,27:(1 974 819)[10.2019] EDI»).
3. Сохраняет в data/stores/all_stores.json с полями:
   id, brand, format, city, address, raw_name.

Парсинг сайтов hitdiscount.by / groshyk.by — НЕ работает: они стоят
за анти-бот challenge (Cloudflare-style verification page). Этот парсер —
основной источник магазинов для RAG.

Использование:
    python3.11 scripts/parse_stores_xlsx.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

DATA_DIR = Path("data")
INPUT_XLSX = DATA_DIR / "stores_new" / "Список ТО ЕТ Хит с форматами.xlsx"
OUTPUT_JSON = DATA_DIR / "stores" / "all_stores.json"


# Регулярки для парсинга адреса из «raw_name».
# Поддерживаемые префиксы населённых пунктов:
#   г.   — город (Минск, Гомель)
#   г.п. — городской посёлок (Бобр, Желудок)
#   п.   — посёлок (Калинино)
#   д.   — деревня (Сорочи, Крево)
#   аг   — агрогородок (Семково, Ратомка) — может быть без точки
#   АГ   — то же, прописью
# Пример: «33 Магазин г.Минск,ул.Рафиева,27:(1 974 819)[10.2019] EDI»
# Пример: «140 АГ д.Сорочи, пер.Школьный, 5А»
# Пример: «190 Магазин аг Ратомка, ул.Минская, 10Б»
RE_CITY = re.compile(
    r"(?:г\.\s*п\.\s*|г\.\s*|п\.\s*|д\.\s*|аг\.?\s+|АГ\s+)"
    r"([А-Яа-яЁё][А-Яа-яЁё\-\s]*?)"
    r"(?=\s*,)",
    re.UNICODE,
)
RE_STREET = re.compile(
    r"(?:ул\.|пр\.|пр-т|просп\.|пер\.|пл\.|бульв\.|б-р|шос\.)\s*([^,]+)",
    re.UNICODE | re.IGNORECASE,
)
RE_HOUSE = re.compile(r",\s*(\d+[А-Яа-я]?(?:\s*корп[\.\s]*\d+)?(?:[\-/\.]?\d*)?)")


# Канонизация названий форматов: внутренний код → читаемая категория
FORMAT_CANONICAL: dict[str, str] = {
    "0 Маркет": "Маркет",
    "1 Маркет": "Маркет",
    "2 Супер": "Супермаркет",
    "3 Евроопт (сухой)": "Магазин (сухой ассортимент)",
    "5 Хит-экспресс (холод)": "Хит-Экспресс",
    "6 Хит Стандарт": "Хит Стандарт",
    "7 Гипер": "Гипермаркет",
    "9 Минимаркет (магазины сельской местности)": "Минимаркет (село)",
    "13 Автолавки": "Автолавка",
    "31 Кафетерий": "Кафетерий",
}

# Канонизация брендов
BRAND_CANONICAL: dict[str, str] = {
    "Евроопт": "Евроопт",
    "Хит": "Хит",
}


def parse_address(raw_name: str) -> dict:
    """Извлечь структурированные поля из строки `33 Магазин г.Минск,ул.Рафиева,27:...`."""
    out = {"city": "", "street": "", "house": "", "address": ""}

    if not raw_name:
        return out

    # Город
    m = RE_CITY.search(raw_name)
    if m:
        out["city"] = m.group(1).strip()

    # Улица
    m = RE_STREET.search(raw_name)
    if m:
        out["street"] = m.group(1).strip().rstrip(",")

    # Дом
    m = RE_HOUSE.search(raw_name)
    if m:
        out["house"] = m.group(1).strip()

    # Полный адрес для отображения — всё, что между «г.» и «:(...)»
    m_full = re.match(r"\d+\s+\S+\s+(.+?)(?:\s*[:(\[]|$)", raw_name)
    if m_full:
        full = m_full.group(1).strip()
        # Чистим хвост вида «:(...)» если попал
        full = re.sub(r"\s*[:(\[].*$", "", full)
        out["address"] = full

    return out


def parse_xlsx(xlsx_path: Path = INPUT_XLSX) -> list[dict]:
    if not xlsx_path.exists():
        print(f"ERROR: файл не найден: {xlsx_path}", file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    stores: list[dict] = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        store_id, raw_name, format_group, format_type, brand_raw = row[:5]

        if not store_id or not raw_name:
            skipped += 1
            continue

        brand = BRAND_CANONICAL.get(brand_raw, brand_raw or "Евроопт")
        format_canon = FORMAT_CANONICAL.get(format_type, format_type or "Магазин")

        addr = parse_address(str(raw_name))

        if not addr["city"]:
            # Не смогли распарсить город — пропускаем (бот лучше промолчит,
            # чем выдаст «магазин г.???»)
            skipped += 1
            continue

        stores.append(
            {
                "id": int(store_id) if isinstance(store_id, (int, float)) else store_id,
                "brand": brand,
                "format": format_canon,
                "format_raw": format_type,
                "format_group": format_group,
                "city": addr["city"],
                "address": addr["address"] or addr["street"],
                "street": addr["street"],
                "house": addr["house"],
                "raw_name": str(raw_name),
            }
        )

    print(f"Загружено магазинов: {len(stores)} (пропущено {skipped})")
    return stores


def stats(stores: list[dict]) -> None:
    """Статистика по брендам/форматам/городам — для проверки."""
    by_brand: dict[str, int] = {}
    by_format: dict[str, int] = {}
    by_city: dict[str, int] = {}
    for s in stores:
        by_brand[s["brand"]] = by_brand.get(s["brand"], 0) + 1
        by_format[s["format"]] = by_format.get(s["format"], 0) + 1
        by_city[s["city"]] = by_city.get(s["city"], 0) + 1

    print("\nПо брендам:")
    for b, n in sorted(by_brand.items(), key=lambda x: -x[1]):
        print(f"  {b:20} {n:>4}")
    print("\nПо форматам:")
    for f, n in sorted(by_format.items(), key=lambda x: -x[1]):
        print(f"  {f:35} {n:>4}")
    print("\nТоп-15 городов:")
    for c, n in sorted(by_city.items(), key=lambda x: -x[1])[:15]:
        print(f"  {c:25} {n:>4}")


def save(stores: list[dict], out_path: Path = OUTPUT_JSON) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(stores, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n→ Сохранено: {out_path} ({len(stores)} магазинов)")


def main() -> int:
    print(f"=== Парсинг {INPUT_XLSX.name} ===\n")
    stores = parse_xlsx()
    if not stores:
        return 1
    stats(stores)
    save(stores)
    print("\nДля загрузки в RAG: python3.11 scripts/reindex_all.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
