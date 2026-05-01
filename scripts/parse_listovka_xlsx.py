"""Парсер xlsx-листовки от заказчика → data/promotions/listovka_current.json.

Заказчик 01.05.2026 прислал спарсенные товары/цены из официальной листовки
evroopt.by/deals (PDF) + страниц «СВАЁ» и «Родныя тавары».

Вход: xlsx с листами «Товары», «Сводка», «Акции», «СВАЁ_ассорт», «Родныя_новости».
Выход: data/promotions/listovka_current.json — структурированный снимок,
который индексируется в RAG через scripts/reindex_v2.py::load_listovka_current().

Запуск:
    python3.11 scripts/parse_listovka_xlsx.py <path_to.xlsx>

По умолчанию ищет последний xlsx в data/promotions_raw/.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
OUT_FP = ROOT / "data/promotions/listovka_current.json"

# ---- helpers ----------------------------------------------------------------

PRICE_RE = re.compile(r"^\s*-?\d+([.,]\d+)?\s*$")


def to_float(x) -> float | None:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace(",", ".").strip()
    if not s or not PRICE_RE.match(s):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def clean_str(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def norm_date(x) -> str:
    """Приводит дату к строке dd.mm.yyyy если возможно."""
    if x is None:
        return ""
    if isinstance(x, datetime):
        return x.strftime("%d.%m.%Y")
    s = str(x).strip()
    return s


# ---- parsing ----------------------------------------------------------------

def parse_products(ws) -> list[dict]:
    """Лист «Товары» / «Исходная частичная»."""
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # Колонки: дата, акция, действует до, название, цена, старая, скидка, за ед.,
        #          категория/прим., статус, URL, источник, исходный фрагмент
        dt = norm_date(row[0])
        promo = clean_str(row[1])
        valid_until = norm_date(row[2])
        name = clean_str(row[3])
        price = to_float(row[4])
        old_price = to_float(row[5]) if len(row) > 5 else None
        discount = clean_str(row[6]) if len(row) > 6 else ""
        per_unit = clean_str(row[7]) if len(row) > 7 else ""
        category = clean_str(row[8]) if len(row) > 8 else ""
        status = clean_str(row[9]) if len(row) > 9 else ""
        url = clean_str(row[10]) if len(row) > 10 else ""
        source = clean_str(row[11]) if len(row) > 11 else ""
        fragment = clean_str(row[12]) if len(row) > 12 else ""

        if not name or price is None:
            continue
        out.append({
            "date": dt,
            "promo": promo,
            "valid_until": valid_until,
            "name": name,
            "price": price,
            "old_price": old_price,
            "discount": discount,
            "per_unit": per_unit,
            "note": category,
            "status": status,
            "url": url,
            "source": source,
            "fragment": fragment,
        })
    return out


def dedup_products(rows: list[dict]) -> list[dict]:
    """Удалить дубликаты по (name, price, promo)."""
    seen = set()
    out = []
    for r in rows:
        key = (r["name"].lower().strip(), round(r["price"], 2), r["promo"].lower().strip())
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def parse_svae_assortment(ws) -> list[dict]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        out.append({
            "code": clean_str(row[0]),
            "name": clean_str(row[1]),
            "category": clean_str(row[2]) if len(row) > 2 else "",
            "url": clean_str(row[3]) if len(row) > 3 else "",
            "note": clean_str(row[4]) if len(row) > 4 else "",
        })
    return out


def parse_rodnyya_news(ws) -> list[dict]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        out.append({
            "date": clean_str(row[0]),
            "type": clean_str(row[1]) if len(row) > 1 else "",
            "title": clean_str(row[2]),
            "products": clean_str(row[3]) if len(row) > 3 else "",
            "url": clean_str(row[4]) if len(row) > 4 else "",
            "fragment": clean_str(row[5]) if len(row) > 5 else "",
        })
    return out


def parse_promo_index(ws) -> list[dict]:
    """Лист «Акции» — индекс акций / разделов."""
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out.append({
            "promo": clean_str(row[0]),
            "pages": clean_str(row[1]) if len(row) > 1 else "",
            "period": clean_str(row[2]) if len(row) > 2 else "",
            "comment": clean_str(row[3]) if len(row) > 3 else "",
        })
    return out


def parse_summary(ws) -> dict:
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out[clean_str(row[0])] = clean_str(row[1]) if len(row) > 1 else ""
    return out


# ---- main -------------------------------------------------------------------

def main():
    # Аргумент — путь к xlsx
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1]).expanduser()
    else:
        # Ищем последний xlsx в data/promotions_raw/
        raw_dir = ROOT / "data/promotions_raw"
        if not raw_dir.exists():
            print(f"❌ Не указан xlsx и нет {raw_dir}")
            sys.exit(1)
        cands = sorted(raw_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not cands:
            print(f"❌ В {raw_dir} нет xlsx-файлов")
            sys.exit(1)
        xlsx_path = cands[0]

    if not xlsx_path.exists():
        print(f"❌ Файл не найден: {xlsx_path}")
        sys.exit(1)

    print(f"📥 Читаю {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ---- Товары (основной + резервный) ----
    products = []
    if "Товары" in wb.sheetnames:
        products.extend(parse_products(wb["Товары"]))
    if "Исходная частичная" in wb.sheetnames:
        products.extend(parse_products(wb["Исходная частичная"]))
    products = dedup_products(products)

    # ---- Доп. данные ----
    svae = parse_svae_assortment(wb["СВАЁ_ассорт"]) if "СВАЁ_ассорт" in wb.sheetnames else []
    rodnyya = parse_rodnyya_news(wb["Родныя_новости"]) if "Родныя_новости" in wb.sheetnames else []
    promo_index = parse_promo_index(wb["Акции"]) if "Акции" in wb.sheetnames else []
    summary = parse_summary(wb["Сводка"]) if "Сводка" in wb.sheetnames else {}

    # ---- Метаданные ----
    promo_breakdown = Counter(p["promo"] for p in products if p["promo"])
    out = {
        "meta": {
            "source_xlsx": xlsx_path.name,
            "fetched_at": datetime.utcnow().isoformat() + "Z",
            "period": summary.get("Период основной акции", ""),
            "dump_date": summary.get("Дата выгрузки/сборки", ""),
            "source_pdf": summary.get("Файл-источник", ""),
            "products_count": len(products),
            "promo_breakdown": dict(promo_breakdown.most_common()),
            "note_prices": summary.get("Примечание по ценам", ""),
        },
        "promo_index": promo_index,
        "products": products,
        "svae_assortment": svae,
        "rodnyya_news": rodnyya,
    }

    OUT_FP.parent.mkdir(parents=True, exist_ok=True)
    OUT_FP.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    print()
    print(f"📊 Период:           {out['meta']['period'] or '—'}")
    print(f"📊 Дата выгрузки:    {out['meta']['dump_date'] or '—'}")
    print(f"📊 Товаров:          {len(products)}")
    print("📊 Разбивка по акциям:")
    for promo, n in promo_breakdown.most_common():
        print(f"     • {promo}: {n}")
    print(f"📊 СВАЁ ассортимент: {len(svae)}")
    print(f"📊 Родныя новости:   {len(rodnyya)}")
    print(f"📊 Promo index:      {len(promo_index)}")
    print()
    print(f"📄 Сохранено в {OUT_FP.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
